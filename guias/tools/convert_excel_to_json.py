import re, json, os
import openpyxl

# ======= CONFIG =======
EXCEL_PATH = r"C:\xampp\htdocs\dashboard\guias\tools\REGLAS.xlsx"
OUT_DIR    = r"C:\xampp\htdocs\dashboard\guias\data"

os.makedirs(OUT_DIR, exist_ok=True)

def pad3(x):
    x = str(x).strip()
    if x.isdigit() and len(x) <= 3:
        return x.zfill(3)
    return x

def extract_cpms(text):
    if text is None:
        return []
    s = str(text)
    # cpms like 99402 ó 99402.03 ó 99402.04, commas, 'o'
    codes = re.findall(r"\d+(?:\.\d+)?", s)
    return [c for c in codes]

def extract_prest(text):
    if text is None:
        return []
    s = str(text)
    codes = re.findall(r"\b\d{1,3}\b", s)
    return [pad3(c) for c in codes]

def parse_rc88(wb):
    """
    RC88: CODIGO PRESTACIONAL | DIAGNOSTICO | DESCRIPCION | CPMS | ACCION | NIVEL
    """
    sheet_name = None
    for n in wb.sheetnames:
        if n.upper().replace("_","") == "RC88":
            sheet_name = n
            break
    if not sheet_name:
        return {}

    ws = wb[sheet_name]

    # find header row that contains "CODIGO PRESTACIONAL" and "CPMS"
    header_row = None
    cols = {}
    for r in range(1, 60):
        row = [ws.cell(r,c).value for c in range(1, 20)]
        row_txt = [str(v).upper().strip() for v in row if isinstance(v,str)]
        if any("CODIGO PREST" in t for t in row_txt) and any("CPMS" == t or "CPMS" in t for t in row_txt):
            header_row = r
            # map columns by exact matches
            for c,v in enumerate(row, start=1):
                if isinstance(v,str):
                    t = v.upper().strip()
                    if "CODIGO PREST" in t: cols["prest"] = c
                    if t == "CPMS" or "CPMS" in t: cols["cpms"] = c
                    if "DIAGN" in t: cols["dx"] = c
                    if "DESCRIP" in t: cols["desc"] = c
                    if "ACCI" in t: cols["accion"] = c
                    if "NIVEL" in t: cols["nivel"] = c
            break

    if not header_row or "prest" not in cols or "cpms" not in cols:
        return {}

    rules = {}  # prest -> {"cpms_allowed": set(), "dx_hint": set()}
    for r in range(header_row+1, ws.max_row+1):
        prest_cell = ws.cell(r, cols["prest"]).value
        cpms_cell  = ws.cell(r, cols["cpms"]).value
        dx_cell    = ws.cell(r, cols.get("dx", 999)).value if "dx" in cols else None

        # ignore empty rows
        if prest_cell is None and cpms_cell is None and dx_cell is None:
            continue

        prests = extract_prest(prest_cell)
        cpms   = extract_cpms(cpms_cell)

        # dx sometimes says "No aplica"
        dxs = []
        if dx_cell and isinstance(dx_cell,str):
            dxs = re.findall(r"[A-Z]\d{2,4}[A-Z0-9\.]*", dx_cell.upper())

        if not prests and not cpms:
            continue

        for p in prests:
            rules.setdefault(p, {"cpms_allowed": set(), "dx_hint": set()})
            for c in cpms:
                rules[p]["cpms_allowed"].add(c)
            for d in dxs:
                rules[p]["dx_hint"].add(d)

    # convert sets to lists
    out = {}
    for p,v in rules.items():
        out[p] = {
            "cpms_allowed": sorted(list(v["cpms_allowed"])),
            "dx_hint": sorted(list(v["dx_hint"]))
        }
    return out

def parse_cpms_catalog(wb):
    """
    Build a CPMS name catalog from sheets that contain CPMS + DENOMINACIÓN/PROCEDIMIENTO.
    We use RC_76 and RC89 (and any other similar).
    """
    catalog = {}  # code -> name
    for name in wb.sheetnames:
        ws = wb[name]
        # scan first 40 rows to find headers
        header = None
        col_cpms = None
        col_name = None
        for r in range(1, 50):
            row = [ws.cell(r,c).value for c in range(1, 20)]
            txt = [str(v).upper().strip() for v in row if isinstance(v,str)]
            if not txt:
                continue
            # case 1: has CPMS + DENOMINACIÓN
            if any(t == "CPMS" or "CPMS" in t for t in txt) and any("DENOMIN" in t or "PROCED" in t for t in txt):
                header = r
                for c,v in enumerate(row, start=1):
                    if isinstance(v,str):
                        t = v.upper().strip()
                        if t == "CPMS" or "CPMS" in t: col_cpms = c
                        if "DENOMIN" in t or "PROCED" in t: col_name = c
                break
        if not header or not col_cpms or not col_name:
            continue

        for r in range(header+1, ws.max_row+1):
            cp = ws.cell(r, col_cpms).value
            nm = ws.cell(r, col_name).value
            if cp is None or nm is None:
                continue
            cp_codes = extract_cpms(cp)
            name_txt = str(nm).strip()
            for c in cp_codes:
                if c not in catalog and name_txt:
                    catalog[c] = name_txt

    return catalog

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    reglas = parse_rc88(wb)
    cpms_catalog = parse_cpms_catalog(wb)

    with open(os.path.join(OUT_DIR, "reglas.json"), "w", encoding="utf-8") as f:
        json.dump(reglas, f, ensure_ascii=False, indent=2)

    with open(os.path.join(OUT_DIR, "cpms_catalog.json"), "w", encoding="utf-8") as f:
        json.dump(cpms_catalog, f, ensure_ascii=False, indent=2)

    print("OK ✅ Generado:")
    print(" - data/reglas.json")
    print(" - data/cpms_catalog.json")

if __name__ == "__main__":
    main()
